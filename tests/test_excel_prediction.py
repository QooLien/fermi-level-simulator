import os
import tempfile
import unittest

from matplotlib.figure import Figure
from openpyxl import Workbook, load_workbook

from app import (MEASUREMENT_DEVICES, VoltageVisualizer, convert_voltage_sheets,
                 parse_voltage_list, predict_measurement_record)


class ExcelPredictionTests(unittest.TestCase):
    def test_parse_voltage_list_accepts_v_suffix_and_commas(self):
        self.assertEqual(parse_voltage_list("0.8v, 0.7V，0.6"), [.8, .7, .6])

    def test_batch_conversion_keeps_vt_and_changes_all_idsat_columns(self):
        with tempfile.TemporaryDirectory() as folder:
            source_path = os.path.join(folder, "source.xlsx")
            output_path = os.path.join(folder, "output.xlsx")
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "0.9v"
            headers = ["Lot/Wafer", "Chip ID"]
            for device in MEASUREMENT_DEVICES:
                headers.extend((f"{device} Vt", f"{device} Idsat"))
            sheet.append(headers)
            row = ["LOT_W01", "CHIP_01"]
            for _device in MEASUREMENT_DEVICES:
                row.extend((.4, 50.0))
            sheet.append(row)
            workbook.save(source_path)
            workbook.close()

            created = convert_voltage_sheets(source_path, "0.9v", [.8, .7], output_path)
            self.assertEqual(created, ["0.80v", "0.70v"])
            output = load_workbook(output_path, data_only=True)
            self.assertAlmostEqual(output["0.80v"]["C2"].value, .4)
            self.assertAlmostEqual(output["0.80v"]["D2"].value, 32.0)
            self.assertAlmostEqual(output["0.70v"]["D2"].value, 18.0)
            for column in (4, 6, 8, 10, 12, 14):
                self.assertAlmostEqual(output["0.80v"].cell(2, column).value, 32.0)
            output.close()

    def test_selected_chip_device_preview_uses_its_own_anchor(self):
        dataset = {"voltage": .9}
        record = {
            "Chip ID": "CHIP_01", "Lot/Wafer": "LOT_W01",
            "PUL Vt": .376087, "PUL Idsat": 45.804,
        }
        result = predict_measurement_record(dataset, record, "PUL", [.8, .7])
        self.assertEqual([round(row["vg"], 3) for row in result["rows"]], [.9, .8, .7])
        self.assertAlmostEqual(result["rows"][1]["idsat"], 29.9873802415)
        self.assertAlmostEqual(result["rows"][2]["idsat"], 17.5082126649)

    def test_selected_chip_preview_draws_one_curve_per_vgs(self):
        dataset = {"voltage": .9}
        record = {"PUL Vt": .376087, "PUL Idsat": 45.804}
        visualizer = VoltageVisualizer.__new__(VoltageVisualizer)
        visualizer.prediction_figure = Figure()
        visualizer.prediction_result = predict_measurement_record(
            dataset, record, "PUL", [.8, .7, .6]
        )
        visualizer.prediction_context = {
            "lot": "LOT_W01", "chip": "CHIP_01", "device": "PUL",
            "vt": .376087, "anchor_voltage": .9,
        }
        visualizer._draw_prediction_curves()
        axis = visualizer.prediction_figure.axes[0]
        self.assertEqual(len(axis.lines), 4)
        self.assertEqual(len(axis.collections), 4)
        self.assertIn("CHIP_01", axis.get_title())


if __name__ == "__main__":
    unittest.main()
